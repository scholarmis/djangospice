from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter        
from openpyxl import Workbook
from .inspection import ModelInspector
from .style import Style
from .utils import convert

class WorkbookEngine:
    """Pure logic for constructing the Excel binary."""
    PASSWORD = "password"
    MAX_ROW = 1048576

    def __init__(self, model, dataset, reference_sheets, **kwargs):
        self.model = model
        self.dataset = dataset
        self.reference_sheets = reference_sheets
        self.hidden_fields = kwargs.get('hidden_fields', [])
        self.export_data = kwargs.get('export_data', False)
        self.protect = kwargs.get('protect', True)

    def build(self) -> Workbook:
        wb = Workbook()
        main_ws = wb.active
        main_ws.title = ModelInspector.get_sheet_name(self.model)

        self._write_to_sheet(main_ws, self.dataset, self.export_data)
        Style.apply_header_style(main_ws)
        
        if self.hidden_fields:
            Style.hide_columns_by_name(main_ws, self.hidden_fields)

        for field_name, (title, ds) in self.reference_sheets.items():
            ref_ws = wb.create_sheet(title)
            self._write_to_sheet(ref_ws, ds, True)
            if self.protect:
                Style.protect_sheet(ref_ws, self.PASSWORD)
            self._apply_validation(main_ws, field_name, title, len(ds))
        return wb

    def _write_to_sheet(self, ws, dataset, include_data):
        for col, header in enumerate(dataset.headers, 1):
            ws.cell(row=1, column=col, value=header)
        if include_data:
            for r_idx, row in enumerate(dataset.dict, 2):
                for c_idx, val in enumerate(row.values(), 1):
                    ws.cell(row=r_idx, column=c_idx, value=convert(val))

    def _apply_validation(self, ws, field_name, ref_title, row_count):
        headers = [str(cell.value).lower() for cell in ws[1]]
        if field_name.lower() in headers:
            col_idx = headers.index(field_name.lower()) + 1
            letter = get_column_letter(col_idx)
            dv = DataValidation(type="list", formula1=f"='{ref_title}'!$B$2:$B${row_count + 1}", allow_blank=True)
            dv.sqref = f"{letter}2:{letter}{self.MAX_ROW}"
            ws.add_data_validation(dv)