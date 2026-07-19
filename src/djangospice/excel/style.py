from typing import List
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class Style:
    """Encapsulates styling and worksheet protection logic."""
    @staticmethod
    def apply_header_style(sheet: Worksheet):
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def hide_columns_by_name(sheet: Worksheet, column_names: List[str]):
        header_map = {cell.value: cell.column for cell in sheet[1] if cell.value}
        for name in column_names:
            if name in header_map:
                col_letter = get_column_letter(header_map[name])
                sheet.column_dimensions[col_letter].hidden = True

    @staticmethod
    def protect_sheet(sheet: Worksheet, password: str):
        sheet.sheet_state = 'hidden'
        sheet.protection.sheet = True
        sheet.protection.set_password(password)

